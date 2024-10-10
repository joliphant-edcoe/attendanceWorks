import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO


# load all the templates once
with open(
    os.path.join(os.path.join("templates", "TK12"), "CalDATT_TK_12_Template.xlsx"), "rb"
) as f:
    template_wb_TK12 = BytesIO(f.read())

with open(
    os.path.join(
        os.path.join("templates", "TK12"), "CalDATT_TK_12_Truancy_Template.xlsx"
    ),
    "rb",
) as f:
    truancy_template_wb_TK12 = BytesIO(f.read())

with open(
    os.path.join(os.path.join("templates", "TK8"), "CalDATT_TK_8_Template.xlsx"), "rb"
) as f:
    template_wb_TK8 = BytesIO(f.read())

with open(
    os.path.join(
        os.path.join("templates", "TK8"), "CalDATT_TK_8_Truancy_Template.xlsx"
    ),
    "rb",
) as f:
    truancy_template_wb_TK8 = BytesIO(f.read())

with open(
    os.path.join(os.path.join("templates", "912"), "CalDATT_9_12_Template.xlsx"), "rb"
) as f:
    template_wb_912 = BytesIO(f.read())

with open(
    os.path.join(
        os.path.join("templates", "912"), "CalDATT_9_12_Truancy_Template.xlsx"
    ),
    "rb",
) as f:
    truancy_template_wb_912 = BytesIO(f.read())


class AttendanceExcelGenerator:
    def __init__(self, attend, truancy, filename, filename_supp, school_type="TK-12"):
        self.adatt = attend
        self.truancy = truancy
        self.filename = filename
        self.filename_supp = filename_supp
        self.school_type = school_type
        self.year = "2024-25"
        self.prior_year_str = (
            str(int(self.year[:4]) - 1) + "-" + str(int(self.year[-2:]) - 1)
        )
        self.two_yr_prior_year_str = (
            str(int(self.year[:4]) - 2) + "-" + str(int(self.year[-2:]) - 2)
        )

        if self.school_type == "TK-8":

            self.num_grades = 10
            self.template_wb = load_workbook(template_wb_TK8)
            self.truancy_template_wb = load_workbook(truancy_template_wb_TK8)

        elif self.school_type == "HS":

            self.num_grades = 4
            self.template_wb = load_workbook(template_wb_912)
            self.truancy_template_wb = load_workbook(truancy_template_wb_912)

        else:
            self.num_grades = 14
            self.template_wb = load_workbook(template_wb_TK12)
            self.truancy_template_wb = load_workbook(truancy_template_wb_TK12)

    def append_data_to_templates(self):

        self._write_df_to_excel(
            self.template_wb["By Grade By Year"],
            self.adatt.bygrade,
            [5],
            [1],
            start_row=6,
        )
        self._write_df_to_excel(
            self.template_wb["By Grade By Year"],
            self.adatt.bygrade_prior,
            [5 + self.num_grades + 2],
            [1],
            start_row=6 + self.num_grades + 2,
            year_str=self.prior_year_str,
        )
        self._write_df_to_excel(
            self.template_wb["By Grade By Year"],
            self.adatt.bygrade_two_yr_prior,
            [5 + 2 * (self.num_grades + 2)],
            [1],
            start_row=6 + 2 * (self.num_grades + 2),
            year_str=self.two_yr_prior_year_str,
        )

        self._write_df_to_excel(
            self.template_wb["By Grade, Overall"],
            self.adatt.bygrade,
            [2] + [self.num_grades + a for a in [11, 11, 34, 34]],
            [4, 2, 8, 2, 8],
        )
        self._write_df_to_excel(
            self.template_wb["By Race, Ethnicity"],
            self.adatt.byrace,
            [2, 18, 18, 39, 39, 60],
            [4, 2, 9, 2, 9, 5],
        )
        self._write_df_to_excel(
            self.template_wb["By Gender"],
            self.adatt.bygender,
            [2, 12, 12, 33, 33],
            [5, 2, 8, 2, 8],
        )
        self._write_df_to_excel(
            self.template_wb["By Race & Gender"],
            self.adatt.byracegender,
            [2, 26, 47],
            [4, 2, 2],
            start_col=3,
        )
        self._write_df_to_excel(
            self.template_wb["By Race & Grade"],
            self.adatt.byracegrade,
            [2] + [6 * self.num_grades + a for a in [8, 32]],
            [4, 2, 3],
            start_col=3,
        )
        self._write_df_to_excel(
            self.template_wb["By Sp Needs Status"],
            self.adatt.byIEP,
            [2, 12, 12],
            [3, 2, 8],
        )
        self._write_df_to_excel(
            self.template_wb["By EL Status"],
            self.adatt.byEngLearner,
            [2, 12, 12],
            [3, 2, 8],
        )
        self._write_df_to_excel(
            self.template_wb["By Lunch Status"],
            self.adatt.byFreeReduced,
            [2, 12, 12],
            [3, 2, 8],
        )
        self._write_df_to_excel(
            self.template_wb["By Zip code"],
            self.adatt.byzipcode.reset_index().iloc[:-1, :].iloc[:20, :],
            [2, 30, 53, 75, 96],
            [4, 6, 6, 6, 6],
            start_col=1,
        )
        self._write_df_to_excel(
            self.template_wb["By Zip code"],
            self.adatt.byzipcode.reset_index().iloc[-1, :].to_frame().T,
            start_col=1,
            start_row=26,
        )
        self._write_df_to_excel(
            self.template_wb["Suspensions in Each School"],
            self.adatt.by_suspension_school.reset_index(),
            [2],
            [8],
            start_col=1,
            start_row=5,
        )

        self._write_df_to_excel(
            self.template_wb["List of students"],
            self.adatt.all_students,
            [2],
            [1],
            start_row=4,
            start_col=1,
        )

        self._write_df_to_excel(
            self.template_wb["School Summary"],
            self.adatt.byschool.reset_index(),
            [2],
            [6],
            start_col=1,
        )

        self.template_wb.save(self.filename)

        self._write_df_to_excel(
            self.truancy_template_wb["List of students"],
            self.truancy.all_students,
            [2],
            [1],
            start_row=4,
            start_col=1,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["Types of Absence"],
            self.truancy.absence_types,
            [2, 11, 11, 34, 34],
            [2, 1, 7, 1, 6],
            start_row=4,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["Absences by School"],
            self.truancy.absence_by_school.reset_index(),
            [2],
            [5],
            start_row=5,
            start_col=1,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["Absences by Gender"],
            self.truancy.absence_by_gender,
            [2, 14, 14, 32, 51],
            [4, 2, 9, 5, 5],
            start_row=4,
            start_col=3,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["Absences by Grade"],
            self.truancy.absence_by_grade,
            [2] + [self.num_grades * 3 + a for a in [8, 27, 47, 70]],
            [5, 4, 4, 5, 5],
            start_row=4,
            start_col=3,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["Absences by Race-Ethnicity"],
            self.truancy.absence_by_race,
            [2, 32, 51, 71, 98],
            [4, 4, 4, 5, 5],
            start_row=4,
            start_col=3,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["By Race-Eth & Gender"],
            self.truancy.absence_by_racegender,
            [2, 25, 46],
            [4, 3, 3],
            start_row=4,
            start_col=4,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["By Race-Eth & Gender"],
            self.truancy.absence_by_racegender_not,
            [68],
            [3],
            start_row=70,
            start_col=4,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["NOTs"],
            self.truancy.part1_notifications,
            [2, 12],
            [4, 4],
            start_row=6,
            start_col=2,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["NOTs"],
            self.truancy.part2_notifications,
            [33, 42, 62],
            [4, 4, 4],
            start_row=36,
            start_col=2,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["NOTs by Grade"],
            self.truancy.part1_grade_notifications,
            [2, 10 + self.num_grades],
            [3, 4],
            start_row=6,
            start_col=2,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["NOTs by Grade"],
            self.truancy.part2_grade_notifications,
            [a + self.num_grades for a in [29, 49]],
            [4, 4],
            start_row=self.num_grade + 32,
            start_col=2,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["NOTs by Grade"],
            self.truancy.part3_grade_notifications,
            [67 + self.num_grades],
            [4],
            start_row=2 * self.num_grades + 57,
            start_col=2,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["NOTs by School"],
            self.truancy.part1_school_notifications.reset_index().iloc[:50, :],
            [2],
            [4],
            start_row=6,
            start_col=1,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["NOTs by School"],
            self.truancy.part2_school_notifications.reset_index().iloc[:50, :],
            [58],
            [3],
            start_row=61,
            start_col=1,
        )

        self._write_df_to_excel(
            self.truancy_template_wb["NOTs by School"],
            self.truancy.part3_school_notifications.reset_index().iloc[:50, :],
            [115],
            [1],
            start_row=118,
            start_col=1,
        )

        self.truancy_template_wb.save(self.filename_supp)

    def _write_df_to_excel(
        self,
        ws,
        data_df,
        year_r_list=[],
        year_c_list=[],
        start_row=6,
        start_col=2,
        year_str=None,
    ):
        if not year_str:
            year_str = self.year

        rows = dataframe_to_rows(data_df, index=False, header=False)

        for r_idx, row in enumerate(rows, start_row):
            for c_idx, value in enumerate(row, start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)

        for r, c in zip(year_r_list, year_c_list):
            ws.cell(row=r, column=c, value=year_str)
